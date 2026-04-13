import pytest
from openpyxl import Workbook

from jin.templates import generate_csv_template, generate_xlsx_template
from jin.uploader import parse_csv_upload, parse_xlsx_upload, parse_xlsx_upload_path, validate_upload_rows


def test_csv_upload_parses_rows() -> None:
    rows = parse_csv_upload(
        b"endpoint,dimension_fields,kpi_fields,tolerance_pct\n/api/sales,retailer,data.RSV,10\n"
    )
    assert rows == [
        {
            "endpoint": "/api/sales",
            "dimension_fields": "retailer",
            "kpi_fields": "data.RSV",
            "tolerance_pct": "10",
        }
    ]


def test_empty_uploads_raise_clear_errors() -> None:
    # Validation happens in validate_upload_rows
    with pytest.raises(ValueError, match="Empty file"):
        validate_upload_rows([], set())

    workbook = Workbook()
    # Empty workbook (just headers, no data)
    from io import BytesIO
    buffer = BytesIO()
    workbook.save(buffer)
    with pytest.raises(ValueError, match="Empty file"):
        parse_xlsx_upload(buffer.getvalue())


def test_validate_upload_rows_normalizes_expected_values() -> None:
    dimensions, kpis, normalized, warnings = validate_upload_rows(
        [
            {
                "endpoint": "/api/sales",
                "dimension_fields": "retailer,period",
                "kpi_fields": "data.RSV,data.units",
                "grain_retailer": "amazon",
                "grain_period": "YTD",
                "expected_data.RSV": "100",
                "expected_data.units": "50",
                "tolerance_pct": "12",
            }
        ],
        {"retailer", "period", "data.RSV", "data.units"},
    )
    assert dimensions == ["retailer", "period"]
    assert kpis == ["data.RSV", "data.units"]
    assert normalized[0]["dimensions"] == {"retailer": "amazon", "period": "YTD"}
    assert normalized[0]["expected"] == {"data.RSV": 100.0, "data.units": 50.0}
    assert warnings == []


def test_validate_upload_rows_po_friendly_format() -> None:
    """Test the new PO-friendly format without meta columns."""
    dimensions, kpis, normalized, warnings = validate_upload_rows(
        [
            {
                "retailer": "amazon",
                "period": "YTD",
                "data.RSV": "100",
                "tolerance_pct": "12",
            }
        ],
        {"retailer", "period", "data.RSV"},
        endpoint="/api/sales",
        dimension_fields=["retailer", "period"],
        kpi_fields=["data.RSV"],
    )
    assert dimensions == ["retailer", "period"]
    assert kpis == ["data.RSV"]
    assert normalized[0]["dimensions"] == {"retailer": "amazon", "period": "YTD"}
    assert normalized[0]["expected"] == {"data.RSV": 100.0}
    assert normalized[0]["endpoint"] == "/api/sales"


def test_validate_upload_rows_rejects_missing_grain_and_expected_columns() -> None:
    with pytest.raises(ValueError, match="Missing required columns: grain_retailer"):
        validate_upload_rows(
            [
                {
                    "endpoint": "/api/sales",
                    "dimension_fields": "retailer",
                    "kpi_fields": "data.RSV",
                    "expected_data.RSV": "100",
                    "tolerance_pct": "10",
                }
            ],
            {"retailer", "data.RSV"},
        )

    with pytest.raises(ValueError, match="Missing required columns in upload: retailer"):
        # PO format but missing a column
        validate_upload_rows(
            [
                {
                    "period": "YTD",
                    "data.RSV": "100",
                }
            ],
            {"retailer", "period", "data.RSV"},
            endpoint="/api/sales",
            dimension_fields=["retailer", "period"],
            kpi_fields=["data.RSV"],
        )


def test_validate_upload_rows_po_format_missing_field_config() -> None:
    """PO format with no server field config raises a clear error."""
    with pytest.raises(ValueError, match="simple template format"):
        validate_upload_rows(
            [{"retailer": "amazon", "data.RSV": "100"}],
            {"retailer", "data.RSV"},
            # No endpoint / dimension_fields / kpi_fields provided
        )


def test_validate_upload_rows_expected_fields_mismatch() -> None:
    """Extra columns should be ignored (with warnings) rather than hard-failing."""
    expected = [
        "endpoint", "dimension_fields", "kpi_fields",
        "grain_retailer", "expected_data.RSV", "tolerance_pct",
    ]
    dim_fields, kpi_fields, normalized, warnings = validate_upload_rows(
        [
            {
                "endpoint": "/api/sales",
                "dimension_fields": "retailer",
                "kpi_fields": "data.RSV",
                "grain_retailer": "amazon",
                "expected_data.RSV": "100",
                "tolerance_pct": "10",
                "extra_col": "bad",  # unexpected column
            }
        ],
        {"retailer", "data.RSV"},
        expected_fields=expected,
    )
    assert dim_fields == ["retailer"]
    assert kpi_fields == ["data.RSV"]
    assert len(normalized) == 1
    assert any("Ignored extra column" in str(w) for w in warnings)


def test_validate_upload_rows_invalid_tolerance() -> None:
    """Rows with a non-numeric tolerance raise ValueError."""
    with pytest.raises(ValueError, match="Invalid tolerance"):
        validate_upload_rows(
            [
                {
                    "endpoint": "/api/sales",
                    "dimension_fields": "retailer",
                    "kpi_fields": "data.RSV",
                    "grain_retailer": "amazon",
                    "expected_data.RSV": "100",
                    "tolerance_pct": "not-a-number",
                }
            ],
            {"retailer", "data.RSV"},
        )


def test_validate_upload_rows_out_of_range_tolerance_defaults_with_warning() -> None:
    """Tolerances outside [1, 100] default to 10 and emit a warning."""
    _, _, _, warnings = validate_upload_rows(
        [
            {
                "endpoint": "/api/sales",
                "dimension_fields": "retailer",
                "kpi_fields": "data.RSV",
                "grain_retailer": "amazon",
                "expected_data.RSV": "100",
                "tolerance_pct": "0.5",  # out of range
            }
        ],
        {"retailer", "data.RSV"},
    )
    assert len(warnings) == 1
    assert "out of range" in warnings[0]


def test_validate_upload_rows_bad_expected_value() -> None:
    """Non-numeric expected values raise ValueError."""
    with pytest.raises(ValueError, match="Expected value must be numeric"):
        validate_upload_rows(
            [
                {
                    "endpoint": "/api/sales",
                    "dimension_fields": "retailer",
                    "kpi_fields": "data.RSV",
                    "grain_retailer": "amazon",
                    "expected_data.RSV": "not-a-number",
                    "tolerance_pct": "10",
                }
            ],
            {"retailer", "data.RSV"},
        )


def test_validate_upload_rows_text_expected_label_is_contextual_not_error() -> None:
    dimensions, kpis, normalized, warnings = validate_upload_rows(
        [
            {
                "endpoint": "/api/revenue/{retailer}",
                "dimension_fields": "retailer",
                "kpi_fields": "data[].label,data[].revenue",
                "grain_retailer": "amazon",
                "expected_data[].label": "current",
                "expected_data[].revenue": "4711.9",
                "tolerance_pct": "10",
            }
        ],
        {"retailer", "data[].label", "data[].revenue"},
    )
    assert dimensions == ["retailer", "data[].label"]
    assert kpis == ["data[].revenue"]
    assert normalized[0]["dimensions"]["retailer"] == "amazon"
    assert normalized[0]["dimensions"]["data[].label"] == "current"
    assert normalized[0]["expected"] == {"data[].revenue": 4711.9}
    assert any("treated as contextual field" in item for item in warnings)


def test_validate_upload_rows_duplicate_grain_warns() -> None:
    """Duplicate grain combinations produce a warning and deduplicate by last row wins."""
    _, _, normalized, warnings = validate_upload_rows(
        [
            {
                "endpoint": "/api/sales",
                "dimension_fields": "retailer",
                "kpi_fields": "data.RSV",
                "grain_retailer": "amazon",
                "expected_data.RSV": "100",
                "tolerance_pct": "10",
            },
            {
                "endpoint": "/api/sales",
                "dimension_fields": "retailer",
                "kpi_fields": "data.RSV",
                "grain_retailer": "amazon",  # same grain key → duplicate
                "expected_data.RSV": "200",
                "tolerance_pct": "10",
            },
        ],
        {"retailer", "data.RSV"},
    )
    assert len(warnings) == 1
    assert "duplicate" in warnings[0].lower()
    # Last row wins
    assert normalized[0]["expected"]["data.RSV"] == 200.0


def test_validate_upload_rows_missing_expected_column() -> None:
    """Missing expected_ column raises ValueError."""
    with pytest.raises(ValueError, match="Missing required columns: expected_data.RSV"):
        validate_upload_rows(
            [
                {
                    "endpoint": "/api/sales",
                    "dimension_fields": "retailer",
                    "kpi_fields": "data.RSV",
                    "grain_retailer": "amazon",
                    # expected_data.RSV deliberately missing
                    "tolerance_pct": "10",
                }
            ],
            {"retailer", "data.RSV"},
        )


def test_validate_upload_rows_defaults_technical_metadata() -> None:
    """Technical metadata fields can be omitted and defaulted."""
    dimensions, kpis, normalized, warnings = validate_upload_rows(
        [
            {
                "region": "US",
                "amount": "150",
                "tolerance_pct": "12",
            }
        ],
        {"region", "api_version", "amount"},
        endpoint="/api/sales",
        dimension_fields=["api_version", "region"],
        kpi_fields=["amount"],
    )
    assert dimensions == ["api_version", "region"]
    assert kpis == ["amount"]
    assert normalized[0]["dimensions"]["api_version"] == "latest"
    assert normalized[0]["dimensions"]["region"] == "US"
    assert normalized[0]["expected"]["amount"] == 150.0
    assert warnings == []


def test_generate_csv_template_excludes_technical_metadata() -> None:
    cols = generate_csv_template("/api/sales", ["region", "api_version"], ["amount"]).decode().splitlines()[0].split(",")
    assert cols == ["region", "amount", "tolerance_pct"]


def test_parse_xlsx_upload_path_reads_data_sheet(tmp_path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["retailer", "amount"])
    worksheet.append(["amazon", 12.0])
    worksheet.append(["walmart", 10.0])
    path = tmp_path / "baseline.xlsx"
    workbook.save(path)

    rows = parse_xlsx_upload_path(str(path))
    assert rows == [
        {"retailer": "amazon", "amount": 12.0},
        {"retailer": "walmart", "amount": 10.0},
    ]


def test_parse_xlsx_upload_path_raises_for_empty_workbook(tmp_path) -> None:
    workbook = Workbook()
    path = tmp_path / "empty.xlsx"
    workbook.save(path)
    with pytest.raises(ValueError, match="Empty file"):
        parse_xlsx_upload_path(str(path))
