from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path
import time
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None
 
from jin.technical_metadata import TECHNICAL_METADATA_FIELDS, default_technical_metadata_value


# Used by /jin/api/v2/debug/build so QA can confirm the running process loaded
# the latest module version (mtime alone can be misleading without a restart).
UPLOADER_IMPORTED_AT_NS = time.time_ns()


# ──────────────────────────────────────────────────────────────────────────────
# Column detection helpers
# ──────────────────────────────────────────────────────────────────────────────

def _is_internal_format(row: dict[str, Any]) -> bool:
    """Return True when the upload uses the legacy internal column layout.

    The legacy format has ``endpoint``, ``dimension_fields``, and ``kpi_fields``
    columns embedded in the sheet.  The new PO-friendly format omits them.
    """
    return "endpoint" in row and "dimension_fields" in row and "kpi_fields" in row


def _validate_rows_internal(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Validate and return rows that use the old internal column format."""
    if not rows:  # pragma: no cover
        raise ValueError("Empty file")
    required = {"endpoint", "dimension_fields", "kpi_fields", "tolerance_pct"}
    missing = required - set(rows[0].keys())
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")
    return rows


import re

def _normalize_header(s: str) -> str:
    """Lowercase and strip non-alphanumeric characters for fuzzy matching."""
    return re.sub(r'[^a-z0-9]', '', s.lower())


def _field_leaf(field: str) -> str:
    return field.replace("[]", "").split(".")[-1]


def _is_technical_metadata_field(field: str) -> bool:
    return field in TECHNICAL_METADATA_FIELDS or _field_leaf(field) in TECHNICAL_METADATA_FIELDS


def _technical_default_for_field(field: str, defaults: dict[str, str]) -> str:
    if field in defaults:
        return defaults[field]
    leaf = _field_leaf(field)
    if leaf in defaults:
        return defaults[leaf]
    return default_technical_metadata_value(leaf)


def _looks_like_time(field: str, sample_value: Any) -> bool:
    """Return True if the field name or its value strongly suggest a time grain."""
    leaf = _field_leaf(field).lower()
    if leaf in {"date", "period", "timestamp", "time", "observed_at", "snapshot", "month", "day", "year"}:
        return True
    
    if not sample_value:
        return False
        
    text = str(sample_value).strip()
    # Simple ISO date check: YYYY-MM-DD
    if re.match(r'^\d{4}-\d{2}-\d{2}', text):
        return True
    # Epoch seconds or millis
    if re.match(r'^\d{10,13}$', text):
        return True
    return False


def _parse_group_cell(text: Any, *, row_index: int) -> dict[str, str]:
    """Parse a PO `Group` cell like `retailer=amazon | market=DE`.

    Keys may be leaf names (e.g. `api_version`) or full Jin field names
    (e.g. `data[].label`). Values are returned as strings.
    """
    if text is None:
        return {}
    raw = str(text).strip()
    if not raw:
        return {}
    parts = [p.strip() for p in raw.split("|") if p.strip()]
    out: dict[str, str] = {}
    for part in parts:
        if "=" not in part:
            raise ValueError(
                f"Invalid Group format at row {row_index}: {part!r}. "
                "Expected `key=value` pairs separated by `|`."
            )
        key, value = part.split("=", 1)
        key = str(key or "").strip()
        value = str(value or "").strip()
        if not key:
            raise ValueError(
                f"Invalid Group format at row {row_index}: {part!r}. "
                "Missing key before `=`."
            )
        out[key] = value
    return out


def _validate_rows_po(
    rows: list[dict[str, Any]],
    dimension_fields: list[str],
    kpi_fields: list[str],
    endpoint: str,
) -> list[dict[str, Any]]:
    """Validate rows from the PO-friendly template format with fuzzy matching.

    The PO template has plain field names (e.g. ``Retailer``) that are matched
    fuzzily against the registry (e.g. ``retailer`` or ``data[].retailer``).
    """
    if not rows:  # pragma: no cover
        raise ValueError("Empty file")

    # Optional alternate PO format: a single `Group` column containing `key=value | key=value`
    # pairs for dimensions. We expand it into pseudo-columns so the usual mapping logic applies.
    headers = list(rows[0].keys())
    group_header: str | None = None
    for header in headers:
        if _normalize_header(str(header or "")) == "group":
            group_header = str(header)
            break
    if group_header is not None:
        expanded_keys: set[str] = set()
        expanded_rows: list[dict[str, Any]] = []
        for idx, row in enumerate(rows, start=2):
            parsed = _parse_group_cell(row.get(group_header), row_index=idx)
            expanded_keys.update(parsed.keys())
            expanded_rows.append({**row, **parsed})
        rows = expanded_rows
        headers = [h for h in list(rows[0].keys()) if str(h) != group_header]
        # Add the expanded keys so header_map can match them.
        headers.extend(sorted(expanded_keys))
    # Create a mapping from normalized header to original header in the file
    header_map = {_normalize_header(h): h for h in headers if h}
    
    # Map required fields to actual headers in the file
    field_to_header: dict[str, str] = {}
    missing = []
    
    all_needed = dimension_fields + kpi_fields
    for field in all_needed:
        # 1. Exact match (case-insensitive)
        exact_match = next((h for h in headers if h.lower() == field.lower()), None)
        if exact_match:
            field_to_header[field] = exact_match
            continue
            
        # 2. Normalized match (strip non-alphanumeric)
        norm_field = _normalize_header(field)
        if norm_field in header_map:
            field_to_header[field] = header_map[norm_field]
            continue
            
        # 3. Leaf match (e.g. "date" matches "data[].date")
        leaf = _field_leaf(field)
        norm_leaf = _normalize_header(leaf)
        if norm_leaf in header_map:
            field_to_header[field] = header_map[norm_leaf]
            continue
            
        # 4. Fuzzy / Alias fallback
        found = False
        aliases = {
            "product": ["sku", "item", "grain"],
            "name": ["sku", "item", "grain"],
            "target": ["expected", "baseline", "ref"],
            "goal": ["expected", "baseline", "ref"],
            "actual": ["value", "reality"],
            "revenue": ["amount", "price", "sales"],
            "sales": ["revenue", "amount", "orders"],
            "stock": ["inventory", "instock", "count"],
        }

        for norm_h, original_h in header_map.items():
            if not norm_h: continue
            
            # Simple substring match
            if norm_h in norm_field or norm_field in norm_h:
                field_to_header[field] = original_h
                found = True
                break
            
            # Alias match
            for alias_key, targets in aliases.items():
                if alias_key in norm_h:
                    if any(t in norm_field for t in targets):
                        field_to_header[field] = original_h
                        found = True
                        break
            if found: break
            
        if not found and not _is_technical_metadata_field(field):
            missing.append(field)

    if missing:
        if group_header is not None:
            raise ValueError(
                "Missing required fields in upload. "
                "If you are using the `Group` column format, include missing dimension fields as "
                "`key=value` pairs in `Group` (or add separate columns). "
                f"Missing: {', '.join(sorted(missing))}. "
                f"Expected fields: {', '.join(sorted(all_needed))}."
            )
        raise ValueError(
            f"Missing required columns in upload: {', '.join(sorted(missing))}. "
            f"Expected columns: {', '.join(sorted(all_needed))}."
        )

    # Inject the internal fields so the rest of the pipeline understands the rows
    out: list[dict[str, Any]] = []
    for row in rows:
        # Order must match expected_upload_columns: 
        # [endpoint, dimension_fields, kpi_fields, *grains, *kpis, tolerance_pct]
        enriched: dict[str, Any] = {
            "endpoint": endpoint,
            "dimension_fields": ",".join(dimension_fields),
            "kpi_fields": ",".join(kpi_fields),
        }
        for field in dimension_fields:
            actual_header = field_to_header.get(field)
            enriched[f"grain_{field}"] = row.get(actual_header, "") if actual_header else ""
        for field in kpi_fields:
            actual_header = field_to_header.get(field)
            enriched[f"expected_{field}"] = row.get(actual_header, "") if actual_header else ""
        
        # Preservation logic: keep ALL original headers in the enriched row
        # This allows validate_upload_rows to find candidates for autonomous promotion
        for header, value in row.items():
            # Don't carry over tolerance columns from PO templates; we normalize them into `tolerance_pct`.
            # Keeping them would break strict template checks later in validation.
            normalized_header = _normalize_header(str(header or ""))
            if normalized_header in {"tolerancepct", "tolerance"}:
                continue
            if header not in field_to_header.values() and header not in enriched:
                # Store them as-is; validate_upload_rows will scan these
                enriched[header] = value

        enriched["tolerance_pct"] = str(row.get("tolerance_pct") or row.get("Tolerance %") or "10")
        out.append(enriched)
    return out


# ──────────────────────────────────────────────────────────────────────────────
# Public helpers (used by the router)
# ──────────────────────────────────────────────────────────────────────────────

def expected_upload_columns(dimension_fields: list[str], kpi_fields: list[str]) -> list[str]:
    return [
        "endpoint",
        "dimension_fields",
        "kpi_fields",
        *[f"grain_{field}" for field in dimension_fields],
        *[f"expected_{field}" for field in kpi_fields],
        "tolerance_pct",
    ]


def parse_csv_upload(content: bytes) -> list[dict[str, Any]]:
    rows = list(csv.DictReader(StringIO(content.decode())))
    return rows  # raw; let validate_upload_rows normalise


def parse_csv_upload_path(path: str) -> list[dict[str, Any]]:
    with Path(path).open("r", encoding="utf-8", newline="") as stream:
        return list(csv.DictReader(stream))


def parse_xlsx_upload(content: bytes) -> list[dict[str, Any]]:
    if load_workbook is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required for XLSX uploads")
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    # Always use first sheet named "Data" if present, otherwise fall back to active
    sheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration as exc:
        raise ValueError("Empty file") from exc
    headers = [str(value) for value in header_row]
    data = [dict(zip(headers, row)) for row in iterator if any(value is not None for value in row)]
    return data  # raw; let validate_upload_rows normalise


def parse_xlsx_upload_path(path: str) -> list[dict[str, Any]]:
    if load_workbook is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required for XLSX uploads")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration as exc:
        raise ValueError("Empty file") from exc
    headers = [str(value) for value in header_row]
    return [dict(zip(headers, row)) for row in iterator if any(value is not None for value in row)]


# ──────────────────────────────────────────────────────────────────────────────
# Core validation + normalisation
# ──────────────────────────────────────────────────────────────────────────────

def validate_upload_rows(
    rows: list[dict[str, Any]],
    field_names: set[str],
    expected_fields: list[str] | None = None,
    *,
    endpoint: str = "",
    dimension_fields: list[str] | None = None,
    kpi_fields: list[str] | None = None,
    technical_defaults: dict[str, str] | None = None,
) -> tuple[list[str], list[str], list[dict[str, Any]], list[str]]:
    """Validate and normalise upload rows.

    Accepts both the new PO-friendly format (plain column names) and the
    legacy internal format (``endpoint``/``dimension_fields``/``kpi_fields``
    meta-columns).

    Returns ``(dimension_fields, kpi_fields, normalised_rows, warnings)``.
    """
    if not rows:
        raise ValueError("Empty file")

    if _is_internal_format(rows[0]):
        rows = _validate_rows_internal(rows)
    else:
        if not dimension_fields or not kpi_fields:
            raise ValueError(
                "The upload file uses the simple template format but "
                "the server did not receive the expected field configuration. "
                "Please download a fresh template and try again."
            )
        rows = _validate_rows_po(rows, dimension_fields, kpi_fields, endpoint or "unknown")

    # Now rows are guaranteed to be in internal format
    raw_dimensions = [item.strip() for item in str(rows[0]["dimension_fields"]).split(",") if item.strip()]
    raw_kpis = [item.strip() for item in str(rows[0]["kpi_fields"]).split(",") if item.strip()]

    # Always initialize warnings before any potential append in the logic below.
    warnings: list[str] = []

    # Validate that configured dims/kpis are real API fields (or match a leaf alias).
    field_leaf_set = {_field_leaf(name) for name in field_names}
    for field in [*raw_dimensions, *raw_kpis]:
        if _is_technical_metadata_field(field):
            continue
        if field in field_names:
            continue
        if _field_leaf(field) in field_leaf_set:
            continue
        raise ValueError(f"Unknown field in upload: {field}")

    # Promotion logic: if we have a field that looks like time but isn't a dimension,
    # and it was provided in the upload, we should promote it to a dimension.
    # This ensures it's part of the grain key and avoids duplicate row warnings.
    # We do this AUTONOMOUSLY even if it's not in field_names, as long as it looks like time.
    raw_dimension_leafs = {_field_leaf(d) for d in raw_dimensions}
    for field in sorted(rows[0].keys()):
        # Strip grain_ prefix if it exists (legacy format usually handles this, but let's be safe)
        clean_field = field[len("grain_"):] if field.startswith("grain_") else field
        
        # Don't promote if it's already a dimension (full match or leaf match) or a KPI
        if clean_field in raw_dimensions or _field_leaf(clean_field) in raw_dimension_leafs or clean_field in raw_kpis:
            continue
            
        # Don't promote if it explicitly looks like a KPI (starts with expected_)
        if field.startswith("expected_"):
            continue

        sample_val = rows[0].get(field)
        if _looks_like_time(clean_field, sample_val):
            # Prefer promoting to the *real API field name* when the upload uses a leaf column
            # like `date` but the API field is `data[].date`. This avoids demanding a confusing
            # `grain_date` column and keeps the grain aligned to API extraction.
            promoted = clean_field
            if clean_field not in field_names:
                leaf = _field_leaf(clean_field)
                matches = sorted({name for name in field_names if _field_leaf(name) == leaf})
                if len(matches) == 1:
                    promoted = matches[0]

            # Promote it (avoid duplicates).
            if promoted not in raw_dimensions and _field_leaf(promoted) not in raw_dimension_leafs:
                raw_dimensions.append(promoted)
                raw_dimension_leafs.add(_field_leaf(promoted))

            # Ensure required `grain_*` column exists so downstream validation passes.
            promoted_col = f"grain_{promoted}"
            for row in rows:
                if promoted_col in row:
                    continue
                # Copy from the original leaf header (e.g. `date`) if present.
                source_value = ""
                if field in row:
                    source_value = row.get(field, "")
                else:
                    leaf = _field_leaf(clean_field)
                    if leaf in row:
                        source_value = row.get(leaf, "")
                row[promoted_col] = source_value

            warnings.append(
                f"Auto-promoted {_field_leaf(promoted)!r} to a dimension as it appears to be the time grain."
            )
            # Note: we don't break here because there might be multiple (e.g. date + period)

    # PO friendliness for "internal" files or manually edited templates:
    # If the file has internal meta-columns but the user provided leaf columns (e.g. `date`)
    # instead of `grain_date`, map them automatically and warn.
    # This prevents confusing errors like "Missing required columns: grain_date" when `date` exists.
    def _backfill_internal_columns() -> None:
        # Build a set of candidate input column names (leafs) available in the file.
        for row_index, row in enumerate(rows, start=2):
            keys = {str(k) for k in row.keys() if k is not None}
            for dim in raw_dimensions:
                target = f"grain_{dim}"
                if target in row:
                    continue
                leaf = _field_leaf(dim)
                # Prefer exact key, then leaf.
                source_key = dim if dim in keys else leaf if leaf in keys else None
                if source_key is None:
                    continue
                row[target] = row.get(source_key, "")
                msg = f"Mapped column {source_key!r} into required template column {target!r} (row {row_index})."
                if msg not in warnings:
                    warnings.append(msg)
            for kpi in raw_kpis:
                target = f"expected_{kpi}"
                if target in row:
                    continue
                leaf = _field_leaf(kpi)
                source_key = kpi if kpi in keys else leaf if leaf in keys else None
                if source_key is None:
                    continue
                row[target] = row.get(source_key, "")
                msg = f"Mapped column {source_key!r} into required template column {target!r} (row {row_index})."
                if msg not in warnings:
                    warnings.append(msg)

    _backfill_internal_columns()

    if expected_fields is not None:
        missing = [column for column in expected_fields if column not in rows[0]]
        if missing:
            raise ValueError(f"Upload columns do not match template: missing {', '.join(missing)}")
        extra = [column for column in rows[0].keys() if column not in expected_fields]
        # Be flexible: ignore extra columns (common when POs add context columns like `label`).
        if extra:
            extra_labels = ", ".join(sorted({str(col) for col in extra if col is not None}))
            if extra_labels:
                warnings.append(
                    f"Ignored extra column(s) not used by Jin: {extra_labels}. "
                    "If this was unintended, download a fresh template for this API."
                )
            # Important: do not delete extra columns here. Some extras (like a plain `date`)
            # may be used by autonomous promotion logic above to backfill required grain_* columns.

    defaults = technical_defaults or {}
    promoted_dimensions: list[str] = []
    promoted_dimension_set: set[str] = set()

    normalized = []
    for index, row in enumerate(rows, start=2):
        tolerance = row.get("tolerance_pct", 10)
        try:
            tolerance_value = float(tolerance)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid tolerance at row {index}") from exc
        if not 1 <= tolerance_value <= 100:
            warnings.append(f"Row {index}: tolerance_pct {tolerance_value} out of range, defaulted to 10")
            tolerance_value = 10.0

        entry: dict[str, Any] = {
            "endpoint": row["endpoint"],
            "dimension_fields": raw_dimensions,
            "kpi_fields": raw_kpis,
            "tolerance_pct": tolerance_value,
            "dimensions": {},
            "expected": {},
        }
        for key, value in row.items():
            if key.startswith("grain_"):
                entry["dimensions"][key.removeprefix("grain_")] = "" if value is None else str(value)
            if key.startswith("expected_"):
                expected_field = key.removeprefix("expected_")
                if value in (None, ""):
                    continue
                try:
                    entry["expected"][expected_field] = float(value)
                except (TypeError, ValueError) as exc:
                    if expected_field in raw_dimensions or _is_technical_metadata_field(expected_field):
                        if expected_field not in promoted_dimension_set and expected_field not in raw_dimensions:
                            promoted_dimensions.append(expected_field)
                            promoted_dimension_set.add(expected_field)
                        entry["dimensions"][expected_field] = str(value)
                        warnings.append(
                            f"Row {index}: {key} is text; treated as contextual field and excluded from numeric KPI checks"
                        )
                        continue
                    raise ValueError(f"Expected value must be numeric at row {index} column {key}") from exc
        for field in raw_dimensions:
            column = f"grain_{field}"
            if column in row and entry["dimensions"].get(field):
                continue
            if _is_technical_metadata_field(field):
                # Only use the default if the column isn't already present in the row
                # (to avoid overwriting user values from PO templates)
                if field not in row and column not in row:
                    entry["dimensions"][field] = _technical_default_for_field(field, defaults)
                continue
            if column not in row:
                raise ValueError(f"Missing required columns: {column}")
        for field in raw_kpis:
            column = f"expected_{field}"
            if column not in row:
                raise ValueError(f"Missing required columns: {column}")
        entry["_row_index"] = index
        normalized.append(entry)

    final_dimensions = [*raw_dimensions, *promoted_dimensions]
    final_kpis = [field for field in raw_kpis if field not in promoted_dimension_set]

    seen_combinations: dict[tuple[str, tuple[tuple[str, str], ...]], int] = {}
    for position, entry in enumerate(normalized):
        for field in final_dimensions:
            if entry["dimensions"].get(field):
                continue
            if _is_technical_metadata_field(field):
                entry["dimensions"][field] = _technical_default_for_field(field, defaults)
                continue
            entry["dimensions"].setdefault(field, "")
        combination_key = (
            str(entry["endpoint"]),
            tuple((field, entry["dimensions"].get(field, "")) for field in final_dimensions),
        )
        row_index = int(entry.get("_row_index", position + 2))
        if combination_key in seen_combinations:
            warnings.append(
                f"Row {row_index}: duplicate grain combination for endpoint {entry['endpoint']}; last row wins"
            )
        seen_combinations[combination_key] = position

    deduped: list[dict[str, Any]] = []
    for index in sorted(seen_combinations.values()):
        item = dict(normalized[index])
        item.pop("_row_index", None)
        deduped.append(item)
    return final_dimensions, final_kpis, deduped, warnings
