from __future__ import annotations

import csv
import json
import re
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from typing import Any

from jin.technical_metadata import TECHNICAL_METADATA_FIELDS
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None


def _is_technical_dimension_field(field: str) -> bool:
    leaf = _field_leaf(field)
    return field in TECHNICAL_METADATA_FIELDS or leaf in TECHNICAL_METADATA_FIELDS


def _business_dimension_fields(dimension_fields: list[str]) -> list[str]:
    return [field for field in dimension_fields if not _is_technical_dimension_field(field)]


def _business_columns(dimension_fields: list[str], kpi_fields: list[str]) -> list[str]:
    """Return the PO-facing column headers for the data sheet.

    Intentionally excludes internal columns like ``endpoint``, ``dimension_fields``,
    ``kpi_fields`` – those are not meaningful to a PO/QA user.
    """
    return [
        *_business_dimension_fields(dimension_fields),
        *kpi_fields,
        "tolerance_pct",
    ]


def _field_leaf(field: str) -> str:
    return field.replace("[]", "").split(".")[-1]


def _field_annotations(fields: list[dict[str, Any]] | None) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for field in fields or []:
        if not isinstance(field, dict):
            continue
        name = str(field.get("name") or "").strip()
        if not name:
            continue
        annotation = str(field.get("annotation") or field.get("type") or "").strip().lower()
        mapping[name] = annotation
        leaf = _field_leaf(name)
        if leaf and leaf not in mapping:
            mapping[leaf] = annotation
    return mapping


def _field_examples(fields: list[dict[str, Any]] | None) -> dict[str, Any]:
    mapping: dict[str, Any] = {}
    for field in fields or []:
        if not isinstance(field, dict):
            continue
        name = str(field.get("name") or "").strip()
        if not name:
            continue
        example = field.get("example")
        if example is None:
            continue
        mapping[name] = example
        leaf = _field_leaf(name)
        if leaf and leaf not in mapping:
            mapping[leaf] = example
    return mapping


def _example_text(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True, separators=(",", ":"))
    return str(value)


def _type_label(annotation: str, field_name: str, metric: bool = False) -> str:
    lower = annotation.lower()
    if lower in {"int", "integer"}:
        return "whole number"
    if lower in {"float", "decimal", "double"}:
        return "decimal number"
    if lower in {"bool", "boolean"}:
        return "true/false"
    if lower in {"date", "datetime"}:
        return "date (YYYY-MM-DD)"
    if metric:
        return "number"
    if re.search(r"(date|time|timestamp|period|snapshot)", field_name.lower()):
        return "date (YYYY-MM-DD)"
    return "text"


def _example_dimension_value(field_name: str, annotation: str, example: Any | None = None) -> str:
    if example is not None:
        return _example_text(example)
    lower = field_name.lower()
    ann = annotation.lower()
    if ann in {"date", "datetime"} or re.search(r"(date|time|timestamp|period|snapshot)", lower):
        return "2026-01-31"
    if re.search(r"(retailer|merchant)", lower):
        return "amazon"
    if re.search(r"(region|zone|area)", lower):
        return "north"
    if re.search(r"(country)", lower):
        return "US"
    if re.search(r"(sku|item|product)", lower):
        return "SKU-1001"
    if re.search(r"(label)", lower):
        return "current"
    if re.search(r"(version)", lower):
        return "v1"
    if re.search(r"(id$|_id$|id_)", lower):
        return "id-001"
    if ann in {"bool", "boolean"}:
        return "true"
    return f"{_field_leaf(field_name)}_a"


def _example_kpi_value(field_name: str, annotation: str, example: Any | None = None) -> str:
    if example is not None:
        return _example_text(example)
    lower = field_name.lower()
    ann = annotation.lower()
    if ann in {"int", "integer"}:
        return "120"
    if ann in {"float", "decimal", "double"}:
        return "0.92"
    if re.search(r"(pct|percent|ratio|rate|through)", lower):
        return "0.92"
    if re.search(r"(count|qty|quantity|units|in_stock|stock|orders)", lower):
        return "120"
    if re.search(r"(revenue|sales|amount|value|price|cost|rsv)", lower):
        return "1500.50"
    return "100"


def _looks_iso_date(value: str) -> bool:
    return bool(re.match(r"^\d{4}-\d{2}-\d{2}$", value))


def _vary_dimension_example(field_name: str, base_value: str, row_index: int) -> str:
    if row_index == 0:
        return base_value
    lower = field_name.lower()

    if _looks_iso_date(base_value):
        try:
            date_value = datetime.strptime(base_value, "%Y-%m-%d").date()
            return (date_value + timedelta(days=7 * row_index)).isoformat()
        except ValueError:
            pass

    if re.search(r"(retailer|merchant)", lower):
        options = ["amazon", "shopify", "walmart"]
        return options[min(row_index, len(options) - 1)]
    if re.search(r"(region|zone|area)", lower):
        options = ["north", "south", "west"]
        return options[min(row_index, len(options) - 1)]
    if re.search(r"(country)", lower):
        options = ["US", "CA", "UK"]
        return options[min(row_index, len(options) - 1)]
    if re.search(r"(period|week)", lower):
        options = ["YTD", "MTD", "QTD"]
        return options[min(row_index, len(options) - 1)]
    if re.search(r"(label)", lower):
        options = ["current", "previous", "target"]
        return options[min(row_index, len(options) - 1)]
    if re.search(r"(version)", lower):
        options = ["v1", "v2", "v3"]
        return options[min(row_index, len(options) - 1)]

    if re.match(r"^id-(\d+)$", base_value):
        try:
            current = int(base_value.split("-")[1])
            return f"id-{current + row_index:03d}"
        except ValueError:
            pass

    if row_index == 1:
        return f"{base_value}_b"
    return f"{base_value}_c"


def _vary_metric_example(base_value: str, row_index: int) -> str:
    if row_index == 0:
        return base_value
    try:
        numeric = float(base_value)
    except (TypeError, ValueError):
        if row_index == 1:
            return f"{base_value}_b"
        return f"{base_value}_c"

    multipliers = [1.0, 1.08, 0.94]
    multiplier = multipliers[min(row_index, len(multipliers) - 1)]
    varied = numeric * multiplier
    if re.match(r"^-?\d+$", str(base_value).strip()):
        return str(int(round(varied)))
    return f"{varied:.2f}"


def _example_row(
    dimension_fields: list[str],
    kpi_fields: list[str],
    fields: list[dict[str, Any]] | None = None,
    *,
    columns: list[str] | None = None,
) -> list[str]:
    annotations = _field_annotations(fields)
    examples = _field_examples(fields)
    dims = set(dimension_fields)
    kpis = set(kpi_fields)
    ordered_columns = list(columns) if columns is not None else _business_columns(dimension_fields, kpi_fields)
    row: list[str] = []
    for column in ordered_columns:
        if column == "tolerance_pct":
            row.append("10")
            continue
        if column in kpis:
            row.append(_example_kpi_value(column, annotations.get(column, ""), examples.get(column)))
            continue
        if column in dims:
            row.append(_example_dimension_value(column, annotations.get(column, ""), examples.get(column)))
            continue
        row.append("")
    return row


def _example_rows(
    dimension_fields: list[str],
    kpi_fields: list[str],
    fields: list[dict[str, Any]] | None = None,
    *,
    columns: list[str] | None = None,
    count: int = 3,
) -> list[list[str]]:
    count = max(1, count)
    base = _example_row(dimension_fields, kpi_fields, fields, columns=columns)
    dims = set(dimension_fields)
    kpis = set(kpi_fields)
    ordered_columns = list(columns) if columns is not None else _business_columns(dimension_fields, kpi_fields)
    rows: list[list[str]] = []
    for row_index in range(count):
        current: list[str] = []
        for col_index, column in enumerate(ordered_columns):
            value = base[col_index] if col_index < len(base) else ""
            if column == "tolerance_pct":
                current.append("10")
            elif column in dims:
                current.append(_vary_dimension_example(column, value, row_index))
            elif column in kpis:
                current.append(_vary_metric_example(value, row_index))
            else:
                current.append(value)
        rows.append(current)
    return rows


def _internal_from_row(
    row: dict[str, str],
    endpoint: str,
    dimension_fields: list[str],
    kpi_fields: list[str],
) -> dict[str, str]:
    """Convert a PO-facing row back to the internal format the uploader expects."""
    out: dict[str, str] = {
        "endpoint": endpoint,
        "dimension_fields": ",".join(dimension_fields),
        "kpi_fields": ",".join(kpi_fields),
        "tolerance_pct": row.get("tolerance_pct", "10"),
    }
    for field in dimension_fields:
        out[f"grain_{field}"] = row.get(field, "")
    for field in kpi_fields:
        out[f"expected_{field}"] = row.get(field, "")
    return out


def template_columns(dimension_fields: list[str], kpi_fields: list[str]) -> list[str]:
    """Internal format columns – kept for backward compat with the parser."""
    return [
        "endpoint",
        "dimension_fields",
        "kpi_fields",
        *[f"grain_{field}" for field in dimension_fields],
        *[f"expected_{field}" for field in kpi_fields],
        "tolerance_pct",
    ]


def generate_csv_template(
    endpoint: str,
    dimension_fields: list[str],
    kpi_fields: list[str],
    fields: list[dict[str, Any]] | None = None,
) -> bytes:
    """Generate a PO-friendly CSV template – no internal columns."""
    buffer = StringIO()
    writer = csv.writer(buffer)
    cols = _business_columns(dimension_fields, kpi_fields)
    writer.writerow(cols)
    writer.writerows(_example_rows(dimension_fields, kpi_fields, fields, columns=cols, count=3))
    return buffer.getvalue().encode()


def generate_xlsx_template(
    endpoint: str,
    dimension_fields: list[str],
    kpi_fields: list[str],
    fields: list[dict[str, Any]] | None = None,
) -> bytes:
    """Generate a PO-friendly XLSX template with a data sheet, instructions sheet,
    and a reference values sheet.

    The data sheet uses plain business column names with no internal Jin columns.
    """
    if Workbook is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required for XLSX template generation")

    wb = Workbook()

    # ─── Sheet 1: Data (the sheet they fill in) ────────────────────────────────
    data_sheet = wb.active
    data_sheet.title = "Data"

    header_fill = PatternFill("solid", fgColor="1E293B") if PatternFill else None
    header_font = Font(bold=True, color="F1F5F9", size=11) if Font else None
    dim_fill = PatternFill("solid", fgColor="0F4C75") if PatternFill else None   # blue tint for group cols
    kpi_fill = PatternFill("solid", fgColor="14532D") if PatternFill else None   # green tint for metric cols
    tol_fill = PatternFill("solid", fgColor="3B1C6E") if PatternFill else None   # purple for tolerance

    cols = _business_columns(dimension_fields, kpi_fields)
    visible_dimensions = _business_dimension_fields(dimension_fields)
    hidden_technical_dimensions = [
        field for field in dimension_fields if _is_technical_dimension_field(field)
    ]
    for col_idx, col_name in enumerate(cols, start=1):
        cell = data_sheet.cell(row=1, column=col_idx, value=col_name)
        if Font:
            cell.font = header_font
        if PatternFill:
            if col_name in visible_dimensions:
                cell.fill = dim_fill
            elif col_name in kpi_fields:
                cell.fill = kpi_fill
            elif col_name == "tolerance_pct":
                cell.fill = tol_fill
            else:  # pragma: no cover
                cell.fill = header_fill
        if Alignment:
            cell.alignment = Alignment(horizontal="center")

    # Example data rows (typed hints derived from API field metadata when available)
    example_rows = _example_rows(dimension_fields, kpi_fields, fields, columns=cols, count=3)
    for row_offset, example in enumerate(example_rows, start=2):
        for col_idx, value in enumerate(example, start=1):
            cell = data_sheet.cell(row=row_offset, column=col_idx, value=value)
            if Font:
                cell.font = Font(color="94A3B8", italic=True, size=10)

    # Column widths
    for col_idx, col_name in enumerate(cols, start=1):
        data_sheet.column_dimensions[_col_letter(col_idx)].width = max(18, len(col_name) + 4)

    data_sheet.freeze_panes = "A2"

    # ─── Sheet 2: How to fill this file ────────────────────────────────────────
    instructions = wb.create_sheet("Instructions")
    instructions.column_dimensions["A"].width = 80

    title_font = Font(bold=True, size=14, color="0F172A") if Font else None
    head_font = Font(bold=True, size=11, color="1E293B") if Font else None
    body_font = Font(size=10, color="334155") if Font else None
    code_font = Font(name="Courier New", size=10, color="7C3AED") if Font else None

    def w(row_num: int, value: str, font=None) -> None:
        cell = instructions.cell(row=row_num, column=1, value=value)
        if font:
            cell.font = font
        if Alignment:
            cell.alignment = Alignment(wrap_text=True)

    r = 1
    w(r, "How to fill this file", title_font); r += 1
    w(r, f"This template is for: {endpoint}"); r += 2

    w(r, "Overview", head_font); r += 1
    w(r, (
        "Fill in one row per combination of group values (the blue columns) and expected metrics (the green columns). "
        "Each row tells Jin what the expected numbers should be for that specific group. "
        "Jin will alert your team when the actual values deviate beyond the tolerance."
    ), body_font); r += 2

    annotations = _field_annotations(fields)
    examples = _field_examples(fields)
    w(r, "Column guide", head_font); r += 1

    if visible_dimensions:
        w(r, "GROUP fields (blue) – Who or what this row is about", Font(bold=True, size=10, color="0F4C75") if Font else None); r += 1
        for field in visible_dimensions:
            example_value = _example_dimension_value(field, annotations.get(field, ""), examples.get(field))
            type_text = _type_label(annotations.get(field, ""), field, metric=False)
            w(r, f"  • {field}  →  {type_text}. Example: {example_value}", body_font); r += 1
        r += 1
    if hidden_technical_dimensions:
        w(
            r,
            "Auto-filled technical fields (not required in upload): "
            + ", ".join(hidden_technical_dimensions),
            body_font,
        ); r += 2

    if kpi_fields:
        w(r, "METRIC fields (green) – Expected numbers for each group", Font(bold=True, size=10, color="14532D") if Font else None); r += 1
        for field in kpi_fields:
            example_value = _example_kpi_value(field, annotations.get(field, ""), examples.get(field))
            type_text = _type_label(annotations.get(field, ""), field, metric=True)
            w(r, f"  • {field}  →  {type_text}. Example: {example_value}", body_font); r += 1
        r += 1

    w(r, "tolerance_pct (purple) – How much variation is acceptable", Font(bold=True, size=10, color="3B1C6E") if Font else None); r += 1
    w(r, (
        "  Enter a number between 1 and 100. This is the allowed % deviation before an alert fires. "
        "Default is 10. For example, 10 means ±10% is fine. Leave it as 10 if unsure."
    ), body_font); r += 2

    w(r, "Rules", head_font); r += 1
    rules = [
        "1. Do not delete or rename the header row.",
        "2. One row = one unique group combination. Duplicate groups will be overwritten by the last row.",
        "3. Metric values must be numbers (no %, no commas, no currency symbols).",
        "4. If you don't know the expected value yet, leave the whole row out – do not enter 0 unless 0 is correct.",
        "5. Do not add extra sheets – only the 'Data' sheet is imported.",
        "6. Save as .xlsx before uploading.",
    ]
    for rule in rules:
        w(r, rule, body_font); r += 1

    r += 1
    w(r, "Example", head_font); r += 1
    if cols:
        example_line = "  " + " | ".join(cols)
        w(r, example_line, code_font); r += 1
        for demo_vals in _example_rows(dimension_fields, kpi_fields, fields, columns=cols, count=3):
            w(r, "  " + " | ".join(demo_vals), code_font); r += 1

    instructions.freeze_panes = None

    # ─── Sheet 3: Reference values (read-only context for the user) ────────────
    ref_sheet = wb.create_sheet("Reference Values")
    ref_sheet.column_dimensions["A"].width = 30
    ref_sheet.column_dimensions["B"].width = 60

    ref_head = Font(bold=True, size=11, color="1E293B") if Font else None
    ref_body = Font(size=10, color="334155") if Font else None

    ref_rows = [
        ("This sheet", "Provided for your reference only. Do not edit it."),
        ("Endpoint", endpoint),
        ("Group fields", ", ".join(visible_dimensions) or "(none configured)"),
        (
            "Auto-filled technical fields",
            ", ".join(hidden_technical_dimensions) or "(none)",
        ),
        ("Metric fields", ", ".join(kpi_fields) or "(none configured)"),
        ("tolerance_pct default", "10  (meaning ±10% before an alert is raised)"),
        ("", ""),
        ("Need help?", "Contact the team who set up this monitoring in Jin."),
    ]
    for row_idx, (key, value) in enumerate(ref_rows, start=1):
        ka = ref_sheet.cell(row=row_idx, column=1, value=key)
        va = ref_sheet.cell(row=row_idx, column=2, value=value)
        if ref_head and key:
            ka.font = ref_head
        if ref_body:
            va.font = ref_body

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _col_letter(n: int) -> str:
    """Convert 1-based column index to Excel letter (A, B, … Z, AA, …)."""
    result = ""
    while n:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def convert_po_rows_to_internal(
    rows: list[dict],
    endpoint: str,
    dimension_fields: list[str],
    kpi_fields: list[str],
) -> list[dict]:
    """Convert PO-friendly rows (from the new template) to the internal format
    expected by ``validate_upload_rows``."""
    return [_internal_from_row(row, endpoint, dimension_fields, kpi_fields) for row in rows]
